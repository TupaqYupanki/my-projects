import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook

#Читаем книгу
wb = load_workbook('book.xlsx')

#делаем активной 2ую страницу
wb.active = 0
ws = wb.active

#ws.cell(row=4, column=3).value - зачение ячейки

y = 1
for row in ws.iter_rows(min_row=1, min_col=3, max_col=3, max_row=200, values_only=True):
    for value in row:
        if 'https' in str(value):
          piccolo = str(value)
          print(piccolo)
          url = requests.get(piccolo)
          if '200' in str(url):
              soup = BeautifulSoup(url.content, features="html.parser")
              mydivs = soup.find_all("span", class_= "ty-price-num")
              if mydivs != []:
                  ws.cell(row=y, column=20).value = mydivs[0].text
              else: ws.cell(row=y, column=20).value = 'price not found'
          else:
              print('url unreachable '+piccolo)
              ws.cell(row=y, column=20).value = 'url unreachable'
    y=y+1
wb.save('test.xlsx')

z = 1
for row in ws.iter_rows(min_row=1, min_col=5, max_col=5, max_row=200, values_only=True):
    for value in row:
        if 'konfigurator' in str(value): print('konfigurator founded')
        elif 'https' in str(value):
            lapsi = str(value)
            print(lapsi)
            url = requests.get(lapsi, headers={'User-Agent':'Test'})
            if '200' in str(url):
                soup = BeautifulSoup(url.content, features="html.parser")
                mydivs = soup.find_all("div", class_= "price")[0]
                if mydivs != []:
                    ws.cell(row=z, column=21).value = mydivs.contents[0].text
                else: ws.cell(row=z, column=21).value = 'price not found'
            else:
                print('url unreachable '+lapsi)
                ws.cell(row=z, column=21).value = 'url unreachable'
    z=z+1

wb.save('test.xlsx')

t = 1
for row in ws.iter_rows(min_row=1, min_col=9, max_col=9, max_row=200, values_only=True):
    for value in row:
        if 'https' in str(value):
            olant = str(value)
            print(olant)
            url = requests.get(olant)
            if '200' in str(url):
                soup = BeautifulSoup(url.content, features="html.parser")
                mydivs = soup.find("div", class_= "card-price__current")
                if mydivs != []:
                    ws.cell(row=t, column=22).value = str(mydivs.text)[:-5]
                else: ws.cell(row=t, column=22).value = 'price not found'
            else:
                print('url unreachable '+olant)
                ws.cell(row=t, column=22).value = 'url unreachable'
    t=t+1

wb.save('test.xlsx')

a = 1
for row in ws.iter_rows(min_row=1, min_col=13, max_col=13, max_row=200, values_only=True):
    for value in row:
        if 'konstruktor' in str(value): print('konstruktor founded')
        elif 'https' in str(value):
            akusher = str(value)
            print(akusher)
            akusherId = akusher.split('/')[4].split('-')[0]
            url = requests.get(akusher)
            if '200' in str(url):
                soup = BeautifulSoup(url.content, features="html.parser")
                mydivs = soup.find("span", id="tovar_price_"+str(akusherId))
                if mydivs != []:
                    ws.cell(row=a, column=23).value = str(mydivs.text)
                else: ws.cell(row=a, column=23).value = 'price not found'
            else:
                print('url unreachable '+akusher)
                ws.cell(row=a, column=23).value = 'url unreachable'
    a=a+1

wb.save('test.xlsx')
