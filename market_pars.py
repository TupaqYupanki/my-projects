import json
from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook('ragnarok1.xlsx')
wb.active = 0
ws = wb.active

with open('market.json') as file:
    market = json.load(file)

with open('items.json') as file:
    items = json.load(file)

my_list = [4064, 4092, 4058, 4105, 4337, 4049, 4126, 4065, 4007, 4040, 4314, 4118, 4141, 4160, 4334, 4279, 4300, 4133, 4077, 4044, 4099, 4111, 4140, 4253, 4249, 4245, 2115, 2421, 2357, 2524, 1230, 2523, 509, 510]
my_items = {}

name = ''
pric = 0

for id in my_list:
    for it in items['items']:
        if it['item_id'] == id:
            print(it['item_id'])
            # name = it['unique_name']
            name = it['name']
            print (name)
            # my_items[it['unique_name']] = {}
            my_items[it['name']] = {}
    for shop in market['shops']:
        for item in shop['items']:
            if item['item_id'] == id:
                print(item)
                # print(item['price'])
                my_items[name][pric] = item['price']
                pric = pric+1
    pric = 0

def full_sorted(name):
    for i in my_items:
        if i == name:
            print(i)
            tmp = sorted(my_items[i].values())
        # print (tmp)
            for x in tmp:
                print(x)

def min_price(row, col):
        row = row
        col = col
        for i in my_items:
            tmp = sorted(my_items[i].values())
            if len(tmp) > 0:
                print(i, tmp[0], len(tmp))
                ws.cell(row=row, column=col).value = i
                # row = row+1
                ws.cell(row=row, column=col).value = tmp[0]
                col = col+1
        wb.save('ragnarok1.xlsx')

min_price(11,1)

# full_sorted('Evil_Druid_Card')




# print(my_items)

# my_items = {itid: 4064, {min_price: 10000, cur_price: 12312}}

#4064 - zerom
#4092 - skelworker
#4058 - thara
#4105 - marc
#4337 - porcello
#4049 - vadon
#4126 - Minouros
#4065 - Kaho
#4007 - Pecopeco Egg
#4040 - Creamy
#4314 - Penomena
#4118 - Earth Petite
#4141 - Evil Druid
#4160 - Firelock Soldier
#4334 - Noxious
#4279 - Earth Deleter
#4300 - Chimera
#4133 - Raydric
#4077 - Phen
#4044 - Smokie
#4099 - Pasana
#4111 - Strouf
#4140 - Abyssal Knight
#4253 - Alice
#4249 - Ancient Worm
#4245 - Am Mut

#2115 - valk shield
#2421 - valk shoes
#2357 - valk armor
#2524 - valk manteaue
#1230 - ice pick
#2523 - undershirt[1]
#509 - white herb
#510 - blue herb
